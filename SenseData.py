from time import sleep
import requests
import pandas as pd
from datetime import datetime, timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import base64

base_url = "https://rp.scrdairy.com/ReverseProxy/rest/api/"
login_api = "v4/auth/login"
herd_api = "animals?offset=0&type=full&includeFilterMetaData=true&isRefresh=true"
daily_health_api = "animals/%d/graphs/2?projection=flat&resolution=day&series=youngStockHealthIndex,dailyRumination,dailyEating,rawRumination,rawEating,rawSuckling,activityTrend"
hourly_health_api = "animals/%d/graphs/2?projection=flat&resolution=hour&series=youngStockHealthIndex,dailyRumination,dailyEating,rawRumination,rawEating,rawSuckling,activityTrend"
daily_heat_api = "animals/%d/graphs/1?projection=flat&resolution=day&series=heatTrend,dailyRumination,dailyEating,activityTrend"
hourly_heat_api = "animals/%d/graphs/1?projection=flat&resolution=hour&series=heatTrend,dailyRumination,dailyEating,activityTrend"

username = input("username: ")
password = input("password: ")
farm = input("farmID: ")


def is_float(element: any) -> bool:
    if element is None:
        return False
    try:
        float(element)
        return True
    except ValueError:
        return False


def login(username: str, password: str, farm: str) -> str:
    auth = f"{username}:{password}"
    auth = base64.b64encode(bytes(auth, "utf-8")).decode("utf-8")
    headers = {
        "Authorization": "Basic " + auth,
        "Farmid": farm,
        "Content-Type": "application/json",
    }
    body = {"username": username, "password": password}
    response = requests.post(base_url + login_api, headers=headers, json=body)
    if response.status_code != 200:
        print("Error:", response.status_code)
    data = response.json()
    return data["result"]["accessToken"]


def herd_list(token: str) -> map:
    headers = {
        "Authorization": "Bearer " + token,
        "Farmid": farm,
    }
    response = requests.get(base_url + herd_api, headers=headers)
    if response.status_code != 200:
        print("Error:", response.status_code)
        return None
    data = response.json()
    rows = data.get("result", {}).get("rows", [])
    map = {}
    for entry in rows:
        map[entry["AnimalIDCalculation"]] = entry["CowDatabaseIDCalculation"]
    return map


def get_data(token: str, id: int) -> dict:
    headers = {
        "Authorization": "Bearer " + token,
        "Farmid": farm,
    }
    response = requests.get(base_url + daily_health_api % id, headers=headers)
    if response.status_code != 200:
        print("Error:", response.status_code)
    return response.json()


def make_excel(data: dict):
    series = data.get("result", {}).get("series", [])

    filtered_data = []
    for entry in series:
        if any(value is not None for key, value in entry.items() if key != "x"):
            readable_date = datetime.fromtimestamp(
                entry["x"], tz=timezone.utc
            ).strftime("%Y-%m-%d %H:%M:%S")
            entry["x"] = readable_date
            filtered_data.append(entry)

    df = pd.DataFrame(filtered_data)

    temp_filename = "temp_output.xlsx"
    df.to_excel(temp_filename, index=False)

    wb = load_workbook(temp_filename)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                if is_float(cell.value):
                    cell.value = float(cell.value)
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output_filename = "%s.xlsx" % vacca
    wb.save(output_filename)
    Path(temp_filename).unlink(missing_ok=True)
    print(f"Created excel file: {output_filename}")


def get_single_data(token: str, herd: map, vacca: str) -> dict:
    data = get_data(token, herd[vacca])
    make_excel(data)


def get_all_herd(token: str, herd: map) -> dict:
    for vacca in herd.items():
        get_single_data(token, herd, vacca)


if __name__ == "__main__":
    token = login(username, password, farm)
    herd = herd_list(token)
    res = input("Do you want the data for all the herd?(Y/N): ")
    if res.lower() == "y":
        get_all_herd(token, herd)
        exit()
    if res.lower() != "n":
        print("Invalid input")
        exit()
    while True:
        vacca = input("Insert the cow ID: ")
        if vacca in herd:
            get_single_data(token, herd, vacca)
            res = input("Do you want to get data for another cow?(Y/N): ")
            if res.lower() == "n":
                break
            if res.lower() != "y":
                print("Invalid input")
                break
